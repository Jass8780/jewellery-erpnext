import frappe
from frappe import _
from frappe.model.mapper import get_mapped_doc
from jewellery_erpnext.jewellery_erpnext.customization.sales_order.doc_events.branch_utils import (
	create_branch_so,
)
from jewellery_erpnext.jewellery_erpnext.doc_events.bom_utils import (
	calculate_gst_rate,
	set_bom_item_details,
	set_bom_rate,
)


def validate(self, method):
	validate_sales_type(self)
	# update_snc(self)
	# update_same_customer_snc(self)
	validate_quotation_item(self)
	# validate_customer_approval_invoice_items(self)
	# if self.sales_type != 'Branch Sales':
	create_new_bom(self)
	validate_serial_number(self)
	# validate_items(self)
	validate_item_dharm(self)
	# calculate_gst_rate(self)
	if not self.get("__islocal") and self.docstatus == 0:
		set_bom_item_details(self)


def on_submit(self, method):
	# submit_bom(self)
	# create_branch_so(self)
	validate_snc(self)


def on_cancel(self, method):
	cancel_bom(self)
	validate_snc(self)



def create_new_bom(self):
	"""
	Optimized: Prefetch data and use dictionary mapping to avoid N+1 queries.
	Creates Sales Order Type BOM from Quotation Bom
	"""
	if not self.items:
		return

	self.total = 0
	
	# 1. Prefetch Global/Static Settings
	settings = frappe.get_cached_doc("Jewellery Settings")
	gold_gst_rate = flt(settings.gold_gst_rate)
	
	# Prefetch Exchange Rate once
	exchange_rate = frappe.cache().get_value("current_exchange_rate_selling")
	if exchange_rate is None:
		exchange_rate_res = frappe.db.sql("""SELECT exchange_rate FROM `tabCurrency Exchange` WHERE for_selling = 1 
			ORDER BY modified DESC LIMIT 1""", pluck="exchange_rate")
		exchange_rate = flt(exchange_rate_res[0]) if exchange_rate_res else 1.0
		frappe.cache().set_value("current_exchange_rate_selling", exchange_rate, expires_in_sec=3600)
	else:
		exchange_rate = flt(exchange_rate)

	# 2. Collect IDs for Batch Queries
	serial_nos = [row.serial_no for row in self.items if row.serial_no]
	
	# 3. Batch Fetch Related Metadata
	serial_metadata = {}
	if serial_nos:
		sn_data = frappe.db.sql("""
			SELECT sn.name, sn.purchase_document_no, se.custom_serial_number_creator as snc,
				snc_doc.parent_manufacturing_order as pmo, pmo_doc.ref_customer, pmo_doc.sales_order
			FROM `tabSerial No` sn
			LEFT JOIN `tabStock Entry` se ON sn.purchase_document_no = se.name
			LEFT JOIN `tabSerial Number Creator` snc_doc ON se.custom_serial_number_creator = snc_doc.name
			LEFT JOIN `tabParent Manufacturing Order` pmo_doc ON snc_doc.parent_manufacturing_order = pmo_doc.name
			WHERE sn.name IN %s
		""", (serial_nos,), as_dict=1)
		
		# Build a lookup map
		for d in sn_data:
			ref_cust = d.ref_customer
			if not ref_cust and d.sales_order:
				ref_cust = frappe.db.get_value("Sales Order", d.sales_order, "ref_customer")
			
			cust_currency = frappe.cache().get_value(f"cust_currency:{ref_cust}") if ref_cust else None
			if ref_cust and not cust_currency:
				cust_currency = frappe.db.get_value("Customer", ref_cust, "default_currency")
				frappe.cache().set_value(f"cust_currency:{ref_cust}", cust_currency, expires_in_sec=3600)
			
			serial_metadata[d.name] = {
				"ref_customer": ref_cust,
				"billing_currency": cust_currency
			}

	# 3. Batch Fetch All BOMs and Settings
	bom_names = list(set([r.bom or r.quotation_bom for r in self.items if (r.bom or r.quotation_bom)]))
	boms_cache = {b.name: b for b in [frappe.get_doc("BOM", n) for n in bom_names]} if bom_names else {}
	
	customer_info = frappe.get_cached_value("Customer", self.customer, ["customer_group", "custom_precision_variable", "custom_gemstone_price_list_type"], as_dict=1)
	precision = cint(customer_info.custom_precision_variable) or 2

	# 4. Iterate and Process
	for row in self.items:
		sn_meta = serial_metadata.get(row.serial_no, {})
		billing_currency = sn_meta.get("billing_currency")
		
		bom_name = row.bom or row.quotation_bom
		if not bom_name or bom_name not in boms_cache: continue
		doc = boms_cache[bom_name]
		
		# Skip branch sales logic if not applicable
		if not row.quotation_bom:
			if self.sales_type != 'Branch Sales': create_serial_no_bom(self, row)
			
			# Optimized calculations
			doc.metal_and_finding_weight = round(sum(flt(r.quantity) for r in doc.metal_detail), precision) + \
										  round(sum(flt(r.quantity) for r in doc.finding_detail), precision)
			
			for gem in getattr(doc, "gemstone_detail", []):
				if self.company == 'Gurukrupa Export Private Limited' and customer_info.customer_group == 'Internal':
					gem.total_gemstone_rate = gem.fg_purchase_rate
				elif self.company == 'KG GK Jewellers Private Limited' and customer_info.customer_group == 'Internal':
					gem.total_gemstone_rate = flt(gem.se_rate) * (exchange_rate if billing_currency == 'USD' else 1.0)
				else:
					# This part could still use a bulk gpc fetch, but for now we'll rely on index-backed get_all
					gpc = frappe.get_all("Gemstone Price List", filters={
						"price_list_type": customer_info.custom_gemstone_price_list_type,
						"gemstone_grade": gem.get("gemstone_grade"), "cut_or_cab": gem.get("cut_or_cab"),
						"gemstone_type": gem.get("gemstone_type"), "stone_shape": gem.get("stone_shape"),
						"customer": self.customer if customer_info.customer_group != "Retail" else None,
						"is_retail_customer": 1 if customer_info.customer_group == "Retail" else 0
					}, fields=["rate", "outwork_handling_charges_rate"], limit=1)
					if gpc:
						rt = flt(gpc[0].rate if not gem.is_customer_item else gpc[0].outwork_handling_charges_rate)
						gem.total_gemstone_rate = round(rt, 2)
				
				gem.gemstone_rate_for_specified_quantity = flt(gem.total_gemstone_rate) / 100 * flt(gem.quantity)
			
			doc.save(ignore_permissions=True)
			row.amount = doc.total_bom_amount + flt(doc.making_charge) + flt(doc.certification_amount)
			row.rate = row.amount / row.qty if row.qty else 0
			row.gold_bom_rate = doc.gold_bom_amount
			row.diamond_bom_rate = doc.diamond_bom_amount
			self.total += row.amount
		else:
			row.bom = row.quotation_bom
			frappe.db.set_value("BOM", row.bom, {"bom_type": "Sales Order", "custom_creation_doctype": "Sales Order", "custom_creation_docname": self.name})
			row.gold_bom_rate = doc.gold_bom_amount
			row.diamond_bom_rate = doc.diamond_bom_amount
			row.rate = doc.total_bom_amount

			# create_sales_order_bom(self, row, diamond_grade_data)


def create_serial_no_bom(self, row):
	serial_no_bom = frappe.db.get_value("Serial No", row.serial_no, "custom_bom_no")
	if not serial_no_bom:
		return
	bom_doc = frappe.get_doc("BOM", serial_no_bom)
	# if self.customer != bom_doc.customer:
	doc = frappe.copy_doc(bom_doc)
	doc.customer = self.customer
	doc.gold_rate_with_gst = self.gold_rate_with_gst
	if hasattr(doc, "diamond_detail"):
		for diamond in doc.diamond_detail or []:
			diamond.quality = self.custom_diamond_quality
		# for diamond in doc.diamond_detail:
	doc.save(ignore_permissions=True)
	row.bom = doc.name
	row.bom_no = doc.name


def create_sales_order_bom(self, row, diamond_grade_data):
	doc = frappe.copy_doc(frappe.get_doc("BOM", row.quotation_bom))
	# doc = get_mapped_doc(
	# 	"BOM",
	# 	row.quotation_bom,
	# 	{
	# 		"BOM": {
	# 			"doctype": "BOM",
	# 		}
	# 	},
	# 	ignore_permissions=True,
	# )
	try:
		doc.custom_creation_doctype = self.doctype
		doc.is_default = 0
		doc.is_active = 1
		doc.bom_type = "Sales Order"
		doc.gold_rate_with_gst = self.gold_rate_with_gst
		doc.customer = self.customer
		doc.selling_price_list = self.selling_price_list
		doc.reference_doctype = "Sales Order"
		doc.reference_docname = self.name
		doc.custom_creation_docname = None
		# doc.save(ignore_permissions=True)
		for diamond in doc.diamond_detail:
			if row.diamond_grade:
				diamond.diamond_grade = row.diamond_grade
				diamond.quality=self.custom_diamond_quality
				
			else:
				if not diamond_grade_data.get(row.diamond_quality):
					diamond_grade_data[row.diamond_quality] = frappe.db.get_value(
						"Customer Diamond Grade",
						{"parent": doc.customer, "diamond_quality": row.diamond_quality},
						"diamond_grade_1",
					)

				diamond.diamond_grade = diamond_grade_data.get(row.diamond_quality)
			if row.diamond_quality:
				diamond.quality = row.diamond_quality

		# This Save will Call before_save and validate method in BOM and Rates Will be Calculated as diamond_quality is calculated too
		doc.save(ignore_permissions=True)
		doc.db_set("custom_creation_docname", self.name)
		row.bom = doc.name
		row.gold_bom_rate = doc.gold_bom_amount
		row.diamond_bom_rate = doc.diamond_bom_amount
		row.gemstone_bom_rate = doc.gemstone_bom_amount
		row.other_bom_rate = doc.other_bom_amount
		row.making_charge = doc.making_charge
		row.bom_rate = doc.total_bom_amount
		row.rate = doc.total_bom_amount
		self.total = doc.total_bom_amount
	except Exception as e:
		frappe.logger("utils").exception(e)
		frappe.log_error(
			title=f"Error while creating Sales Order from {row.quotation_bom}", message=str(e)
		)
		frappe.throw(_("Row {0} {1}").format(row.idx, e))

def validate_snc(self):
	for row in self.items:
		if row.serial_no:
			if self.docstatus == 2:  
				frappe.db.set_value("Serial No", row.serial_no, "status", "Active")
			else:
				frappe.db.set_value("Serial No", row.serial_no, "status", "Reserved")

def submit_bom(self):
	for row in self.items:
		if row.bom:
			bom_doc = frappe.get_doc("BOM", row.bom)
			if bom_doc.docstatus == 0:
				bom_doc.submit()
			# frappe.enqueue(enqueue_submit_bom, job_name="Submitting SO BOM", bom=row.bom)


# def enqueue_submit_bom(bom):
# 	bom_doc = frappe.get_doc("BOM", bom)
# 	if bom_doc.docstatus == 0:
# 		bom_doc.submit()


def cancel_bom(self):
	for row in self.items:
		if row.bom:
			bom = frappe.get_doc("BOM", row.bom)
			bom.is_active = 0
			row.bom = ""

def validate_serial_number(self):
	if getattr(self, 'skip_serial_validation', False):
		return
	
	for row in self.items:
		if row.serial_no:
			# serial_nos = [s.strip() for s in row.serial_no.split('\n') if s.strip()]

			# for serial in serial_nos:
			existing = frappe.db.sql("""
				SELECT soi.name, soi.parent
				FROM `tabSales Order Item` soi
				JOIN `tabSales Order` so ON soi.parent = so.name
				WHERE so.docstatus = 1
					AND soi.serial_no = %s
				
			""", (row.serial_no), as_dict=True)
			# if existing:
			# 	so_name = existing[0].parent
			# 	frappe.throw(f"Serial No {row.serial_no} is already used in submitted Sales Order {so_name}.")



import frappe
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from io import BytesIO

@frappe.whitelist()
def xl_preview_sales_order(docname):
    """
    Optimized Excel Preview: Prefetch BOMs and metadata to avoid N+1 calls in spreadsheet generation.
    """
    doc = frappe.get_doc("Sales Order", docname)
    rows_diamond = []

    # 1. Prefetch all unique BOMs used in the SO items
    bom_names = list(set([item.quotation_bom or item.bom for item in doc.items if (item.quotation_bom or item.bom)]))
    
    boms_cache = {}
    if bom_names:
        for b_name in bom_names:
            boms_cache[b_name] = frappe.get_cached_doc("BOM", b_name)

    columns = [
        "Index","Item Code","Serial No","Item Name","Diamond Quality","PCS","Diamond Weight","Average",
        "Total Cts","Grams","Total Diamond Rate","Diamond Amount","Gross Weight","Gemstone Weight",
        "Other Weight","Gold Rate","Net Weight","Gold Amount","Customer Purity","Chain Weight",
        "Chain Amount","Chain Purity","Per Gram MC","Chain MC","Chain Wastage %","Chain Wastage Amount",
        "Jewellery Per Gram MC","Jewellery MC","Gold Wastage %","Jewellery Wastage","Gemstone Pcs",
        "Gemstone Cts","Gemstone Amount","Cert Charge","Hallmark Charge","Total Amt"
    ]

    # --- Populate rows_diamond ---
    for item in doc.items:
        bom_name = item.quotation_bom or item.bom
        if not bom_name or bom_name not in boms_cache:
            continue

        bom_doc = boms_cache[bom_name]

        total_qty = sum([flt(d.quantity) for d in bom_doc.diamond_detail])
        grams = total_qty * 0.2
        gross_weight = round(flt(bom_doc.gross_weight), 2)
        gemstone_weight = flt(bom_doc.total_gemstone_weight_in_gms)
        other_weight = flt(bom_doc.other_weight)
        net_weight = flt(bom_doc.metal_and_finding_weight)

        gemstone_pcs_rows = [flt(g.pcs) for g in bom_doc.gemstone_detail] if bom_doc.gemstone_detail else []
        gemstone_cts_rows = [flt(g.quantity) for g in bom_doc.gemstone_detail] if bom_doc.gemstone_detail else []
        gemstone_amount_rows = [flt(g.gemstone_rate_for_specified_quantity) for g in bom_doc.gemstone_detail] if bom_doc.gemstone_detail else []

        chain_weight_val, chain_mc_val, chain_wastage_val = 0.0, 0.0, 0.0
        chain_weight, chain_amount, chain_mc, chain_wastage, chain_purity = 0, 0, 0, 0, 0
        per_gram_mc, chain_wastage_amount = 0, 0
        net_weight_from_findings = 0.0

        if bom_doc.finding_detail:
            for f in bom_doc.finding_detail:
                qty = flt(f.quantity)
                if f.finding_category and f.finding_category.lower() == "chains":
                    chain_weight_val += qty
                    chain_purity = flt(f.customer_metal_purity)
                    per_gram_mc = flt(f.making_rate)
                    chain_mc_val = flt(f.making_amount)
                    chain_wastage_val = flt(f.wastage_rate)
                else:
                    net_weight_from_findings += qty

        if chain_weight_val > 0:
            chain_weight = chain_weight_val
            quotation_gold_rate = flt(doc.gold_rate)
            chain_amount = (quotation_gold_rate * chain_purity / 100) * chain_weight
            chain_mc = chain_mc_val
            chain_wastage = chain_wastage_val
            chain_wastage_amount = (chain_amount * chain_wastage_val) if chain_wastage_val else 0

        net_weight_display = net_weight + net_weight_from_findings
        if chain_weight > 0:
            net_weight_display -= chain_weight

        if bom_doc.metal_detail:
            customer_metal_purity = flt(bom_doc.metal_detail[0].customer_metal_purity)
            gold_wastage = flt(bom_doc.metal_detail[0].wastage_rate)
            jewellery_per_gram_mc = flt(bom_doc.metal_detail[0].making_rate)
        else:
            customer_metal_purity, gold_wastage, jewellery_per_gram_mc = 0.0, 0, 0

        calculated_gold_rate = flt(f"{(flt(doc.gold_rate) * customer_metal_purity / 100):.2f}")
        cert_charge = flt(bom_doc.certification_amount)
        hallmark_charge = flt(bom_doc.hallmarking_amount)

        for i, diamond in enumerate(bom_doc.diamond_detail):
            pcs = flt(diamond.pcs)
            qty = flt(f"{flt(diamond.quantity):.2f}")
            avg = (qty / pcs) if pcs else 0
            rate = flt(diamond.total_diamond_rate)
            diamond_amount = rate * qty

            gold_amount_val = calculated_gold_rate * net_weight_display if i == 0 else 0
            jewellery_wastage_val = gold_amount_val * (gold_wastage / 100) if i == 0 else 0

            gemstone_pcs_val = gemstone_pcs_rows[i] if i < len(gemstone_pcs_rows) else 0
            gemstone_cts_val = gemstone_cts_rows[i] if i < len(gemstone_cts_rows) else 0
            gemstone_amount_val = gemstone_amount_rows[i] if i < len(gemstone_amount_rows) else 0
            jewellery_mc_val = net_weight_display * jewellery_per_gram_mc if i == 0 else 0

            total_amt = hallmark_charge + cert_charge + jewellery_mc_val + \
                        gemstone_amount_val + gold_amount_val + \
                        jewellery_wastage_val + diamond_amount

            rows_diamond.append([
                item.idx if i == 0 else "",
                item.item_code if i == 0 else "",
                item.serial_no if i == 0 else "",
                item.item_name if i == 0 else "",
                item.diamond_quality,
                pcs, f"{qty:.2f}", f"{avg:.3f}",
                round(total_qty, 2) if (i == 0 and total_qty != 0) else "",
                round(grams, 2) if (i == 0 and grams != 0) else "",
                round(rate, 2), round(diamond_amount, 2),
                round(gross_weight, 2) if (i == 0 and gross_weight != 0) else "",
                round(gemstone_weight, 2) if (i == 0 and gemstone_weight != 0) else "",
                round(other_weight, 2) if (i == 0 and other_weight != 0) else "",
                f"{calculated_gold_rate:.2f}" if i == 0 else "",
                round(net_weight_display, 2) if i == 0 else "",
                f"{gold_amount_val:.2f}" if i == 0 else "",
                customer_metal_purity if i == 0 else "",
                round(chain_weight, 2) if i == 0 else "",
                round(chain_amount, 2) if i == 0 else "",
                chain_purity if i == 0 else "",
                round(per_gram_mc, 2) if i == 0 else "",
                round(chain_mc, 2) if i == 0 else "",
                round(chain_wastage, 2) if i == 0 else "",
                round(chain_wastage_amount, 2) if i == 0 else "",
                round(jewellery_per_gram_mc, 2) if i == 0 else "",
                round(jewellery_mc_val, 2) if i == 0 else "",
                round(gold_wastage, 2) if i == 0 else "",
                round(jewellery_wastage_val, 2) if i == 0 else "",
                gemstone_pcs_val if gemstone_pcs_val != 0 else "",
                round(gemstone_cts_val, 2) if gemstone_cts_val != 0 else "",
                round(gemstone_amount_val, 2) if gemstone_amount_val != 0 else "",
                round(cert_charge, 2) if i == 0 else "",
                round(hallmark_charge, 2) if i == 0 else "",
                round(total_amt, 2)
            ])


    # --- SUM ROW ---
    sum_row = [""] * len(columns)
    sum_row[5]  = round(sum(float(r[5] or 0) for r in rows_diamond), 2)
    sum_row[6]  = round(sum(float(r[6] or 0) for r in rows_diamond), 2)
    sum_row[8]  = round(sum(float(r[8] or 0) for r in rows_diamond), 2)
    sum_row[10] = round(sum(float(r[10] or 0) for r in rows_diamond), 2)  # Total Diamond Rate
    sum_row[11] = round(sum(float(r[11] or 0) for r in rows_diamond), 2)  # Diamond Amount
    sum_row[12] = round(sum(float(r[12] or 0) for r in rows_diamond), 2)
    sum_row[13] = round(sum(float(r[13] or 0) for r in rows_diamond), 2)
    sum_row[14] = round(sum(float(r[14] or 0) for r in rows_diamond), 2)
    sum_row[16] = round(sum(float(r[16] or 0) for r in rows_diamond), 2)
    sum_row[17] = round(sum(float(r[17] or 0) for r in rows_diamond), 2)
    sum_row[19] = round(sum(float(r[19] or 0) for r in rows_diamond), 2)
    sum_row[20] = round(sum(float(r[20] or 0) for r in rows_diamond), 2)
    sum_row[23] = round(sum(float(r[23] or 0) for r in rows_diamond), 2)
    sum_row[25] = round(sum(float(r[25] or 0) for r in rows_diamond), 2)
    sum_row[27] = round(sum(float(r[27] or 0) for r in rows_diamond), 2)  #  Jewellery MC total
    sum_row[29] = round(sum(float(r[29] or 0) for r in rows_diamond), 2)
    sum_row[30] = round(sum(float(r[30] or 0) for r in rows_diamond), 2)
    sum_row[31] = round(sum(float(r[31] or 0) for r in rows_diamond), 2)
    sum_row[32] = round(sum(float(r[32] or 0) for r in rows_diamond), 2)
    sum_row[33] = round(sum(float(r[33] or 0) for r in rows_diamond), 2)
    sum_row[34] = round(sum(float(r[34] or 0) for r in rows_diamond), 2)
    sum_row[35] = round(sum(float(r[35] or 0) for r in rows_diamond), 2)

    rows_diamond.append(sum_row)

    # --- Create Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Diamond Detail"

    # --- Add Company Name ---
    company_name = "M/S. GURUKRUPA EXPORT PVT LIMITED"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1, value=company_name)
    cell.font = Font(bold=True, size=15)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Add Headers ---
    for col_num, column_title in enumerate(columns, 1):
        c = ws.cell(row=2, column=col_num, value=column_title)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Add Data Rows ---
    for row_num, row_data in enumerate(rows_diamond, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # --- Auto column width ---
    for i, column in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # --- Save to BytesIO and Download ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    frappe.local.response.filecontent = output.read()
    frappe.local.response.filename = f"Diamond_Detail_SO_{docname}.xlsx"
    frappe.local.response.type = "download"


@frappe.whitelist()
def get_customer_approval_data(customer_approval_data):
	doc = frappe.get_doc("Customer Approval", customer_approval_data)
	return doc


@frappe.whitelist()
def customer_approval_filter(doctype, txt, searchfield, start, page_len, filters):
	CustomerApproval = frappe.qb.DocType("Customer Approval")
	StockEntry = frappe.qb.DocType("Stock Entry")

	query = (
		frappe.qb.from_(CustomerApproval)
		.left_join(StockEntry)
		.on(CustomerApproval.name == StockEntry.custom_customer_approval_reference)
		.select(CustomerApproval.name)
		.where(
			(
				(StockEntry.custom_customer_approval_reference != CustomerApproval.name)
				| (StockEntry.custom_customer_approval_reference.isnull())
			)
			& (CustomerApproval.docstatus == 1)
			& (CustomerApproval[searchfield].like(f"%{txt}%"))
		)
	)

	if filters.get("date"):
		query = query.where(CustomerApproval.date == filters["date"])

	dialoge_filter = query.run(as_dict=True)

	return dialoge_filter




from frappe.utils import flt
def validate_items(self):
	# for row in self.custom_invoice_item:
		# frappe.throw(f"{row.rate}")
	allowed = ("Finished Goods", "Subcontracting", "Certification")
	if self.sales_type in allowed:
		customer_payment_term_doc = frappe.get_doc(
			"Customer Payment Terms",
			{"customer": self.customer}
		)
		
		e_invoice_items = []
	
		# Loop through all child table rows
		for row in customer_payment_term_doc.customer_payment_details:
			item_type = row.item_type
			e_invoice_item = frappe.get_doc("E Invoice Item", item_type)
			
			matched_sales_type_row = None
			for row in e_invoice_item.sales_type:
				if row.sales_type == self.sales_type:
					matched_sales_type_row = row
					break

			# Skip item if no matching sales_type and custom_sales_type is set
			if self.sales_type and not matched_sales_type_row:
				continue
			e_invoice_items.append({
				"item_type": item_type,
				"is_for_metal": e_invoice_item.is_for_metal,
				"is_for_labour": e_invoice_item.is_for_labour,
				"is_for_diamond": e_invoice_item.is_for_diamond,
				"diamond_type" : e_invoice_item.diamond_type,
				"is_for_making": e_invoice_item.is_for_making,
				"is_for_finding": e_invoice_item.is_for_finding,
				"is_for_finding_making": e_invoice_item.is_for_finding_making,
				"is_for_gemstone": e_invoice_item.is_for_gemstone,
				"metal_type": e_invoice_item.metal_type,
				"metal_purity": e_invoice_item.metal_purity,
				"uom": e_invoice_item.uom,
				"tax_rate": matched_sales_type_row.tax_rate if matched_sales_type_row else 0
			})
		self.set("custom_invoice_item", [])
		aggregated_metal_items = {}
def validate_items(self):
	"""
	Optimized: Bulk fetch Payment Terms, E-Invoice items, and BOM details.
	"""
	allowed = ("Finished Goods", "Subcontracting", "Certification")
	if self.sales_type not in allowed: return

	# 1. Bulk Fetch Metadata
	term_res = frappe.get_all("Customer Payment Terms", filters={"customer": self.customer}, fields=["name"], limit=1)
	if not term_res: return
	term_doc = frappe.get_cached_doc("Customer Payment Terms", term_res[0].name)
	
	e_types = [r.item_type for r in term_doc.customer_payment_details if r.item_type]
	e_map = {}
	if e_types:
		e_data = frappe.get_all("E Invoice Item", filters={"name": ["in", e_types]}, fields=["*"])
		for e in e_data:
			e.sales_types = frappe.get_all("E Invoice Sales Type", filters={"parent": e.name}, fields=["sales_type", "tax_rate"])
			e_map[e.name] = e

	processed_e = []
	for pd in term_doc.customer_payment_details:
		e = e_map.get(pd.item_type)
		if not e: continue
		match = next((x for x in e.sales_types if x.sales_type == self.sales_type), None)
		if self.sales_type and not match: continue
		processed_e.append({**e, "tax_rate": flt(match.tax_rate) if match else 0})

	# 2. Bulk Fetch BOMs
	bom_ids = list(set([item.bom for item in self.items if item.bom]))
	bom_lookup = {}
	if bom_ids:
		d_rows = frappe.get_all("BOM Diamond Detail", filters={"parent": ["in", bom_ids]}, fields=["*"])
		m_rows = frappe.get_all("BOM Metal Detail", filters={"parent": ["in", bom_ids]}, fields=["*"])
		f_rows = frappe.get_all("BOM Finding Detail", filters={"parent": ["in", bom_ids]}, fields=["*"])
		g_rows = frappe.get_all("BOM Gemstone Detail", filters={"parent": ["in", bom_ids]}, fields=["*"])
		
		from collections import defaultdict
		dm, mm, fm, gm = defaultdict(list), defaultdict(list), defaultdict(list), defaultdict(list)
		for r in d_rows: dm[r.parent].append(r)
		for r in m_rows: mm[r.parent].append(r)
		for r in f_rows: fm[r.parent].append(r)
		for r in g_rows: gm[r.parent].append(r)
		for bid in bom_ids: bom_lookup[bid] = {"diamond": dm[bid], "metal": mm[bid], "finding": fm[bid], "gemstone": gm[bid]}

	# 3. Aggregation
	self.set("custom_invoice_item", [])
	agg = defaultdict(lambda: {"qty": 0, "amount": 0, "tax_amount": 0})
	
	for item in self.items:
		if not item.bom or item.bom not in bom_lookup: continue
		b = bom_lookup[item.bom]
		
		for m in b["metal"]:
			for e in processed_e:
				is_m = e.get("is_for_metal") and m.metal_type == e["metal_type"] and m.metal_touch == e["metal_purity"] and not m.is_customer_item
				is_l = e.get("is_for_labour") and m.is_customer_item
				if is_m or is_l:
					k = (e["name"], e["uom"], e["tax_rate"])
					q = flt(m.quantity) * flt(item.qty)
					r = flt(m.se_rate) if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else flt(m.rate)
					amt = r * q
					agg[k]["qty"] += q; agg[k]["amount"] += amt; agg[k]["tax_amount"] += amt * (e["tax_rate"]/100)
					break

		for d in b["diamond"]:
			if d.is_customer_item: continue
			for e in processed_e:
				if e.get("is_for_diamond") and e["diamond_type"] == d.diamond_type and e["uom"] == d.stock_uom:
					k = (e["name"], e["uom"], e["tax_rate"])
					q = flt(d.quantity) * flt(item.qty)
					amt = flt(d.diamond_rate_for_specified_quantity) # Assuming this is already stored in BOM
					agg[k]["qty"] += q; agg[k]["amount"] += amt; agg[k]["tax_amount"] += amt * (e["tax_rate"]/100)
					break

	for (itype, uom, tax), v in agg.items():
		self.append("custom_invoice_item", {
			"item_code": itype, "item_name": itype, "uom": uom,
			"qty": v["qty"], "amount": v["amount"],
			"rate": v["amount"] / v["qty"] if v["qty"] else 0,
			"tax_rate": tax, "tax_amount": v["tax_amount"],
			"amount_with_tax": v["amount"] + v["tax_amount"],
			"delivery_date": self.delivery_date
		})


def validate_item_dharm(self):
	allowed = ("Finished Goods", "Subcontracting", "Certification","Branch Sales")
	if self.sales_type in allowed:
		customer_payment_term_doc = frappe.get_doc(
			"Customer Payment Terms",
			{"customer": self.customer}
		)
		
		e_invoice_items = []

		for row in self.items:
			gross_weighh = frappe.get_value("BOM", row.bom, "gross_weight")
			row.custom_gross_weight = gross_weighh
			
		# Prepare invoice items as before
		for row in customer_payment_term_doc.customer_payment_details:
			item_type = row.item_type
			e_invoice_item = frappe.get_doc("E Invoice Item", item_type)
			matched_sales_type_row = None
			for st_row in e_invoice_item.sales_type:
				if st_row.sales_type == self.sales_type:
					matched_sales_type_row = st_row
					break

			if self.sales_type and not matched_sales_type_row:
				continue

			e_invoice_items.append({
				"item_type": item_type,
				"is_for_metal": e_invoice_item.is_for_metal,
				"is_for_hallmarking":e_invoice_item.is_for_hallmarking,
				"is_for_labour": e_invoice_item.is_for_labour,
				"is_for_diamond": e_invoice_item.is_for_diamond,
				"diamond_type": e_invoice_item.diamond_type,
				"is_for_making": e_invoice_item.is_for_making,
				"is_for_finding": e_invoice_item.is_for_finding,
				"is_for_finding_making": e_invoice_item.is_for_finding_making,
				"is_for_gemstone": e_invoice_item.is_for_gemstone,
				"metal_type": e_invoice_item.metal_type,
				"metal_purity": e_invoice_item.metal_purity,
				"uom": e_invoice_item.uom,
				"finding_category":e_invoice_item.finding_category,
				"tax_rate": matched_sales_type_row.tax_rate if matched_sales_type_row else 0
			})

		self.set("custom_invoice_item", [])
		aggregated_metal_items = {}
		aggregated_metal_labour_items = {}
		aggregated_metal_making_items = {}
		aggregated_hallmarking_items = {}
		aggregated_diamond_items = {}
		aggregated_gemstone_items = {}
		aggregated_finding_items = {}
		aggregated_finding_making_items = {}
		for item in self.items:
			if item.bom:
				bom_doc = frappe.get_doc("BOM", item.bom)
				if bom_doc.hallmarking_amount:
					for e_item in e_invoice_items:
						if (
							e_item["is_for_hallmarking"]
						):
							key = (e_item["item_type"], e_item["uom"])
							if key not in aggregated_hallmarking_items:
								aggregated_hallmarking_items[key] = {
									"item_code": e_item["item_type"],
									"item_name": e_item["item_type"],
									"uom": e_item["uom"],
									"qty": 0,
									"amount": 0,
									"tax_rate": e_item["tax_rate"],
									"tax_amount": 0,
									"amount_with_tax": 0,
									"delivery_date": self.delivery_date
								}
							aggregated_hallmarking_items[key]["amount"] += bom_doc.hallmarking_amount
							aggregated_hallmarking_items[key]["qty"] +=1

				for metal in bom_doc.metal_detail:
					
					if not metal.is_customer_item:
						for e_item in e_invoice_items:
							if (
								e_item["is_for_metal"] and
								metal.metal_type == e_item["metal_type"] and
								metal.metal_touch == e_item["metal_purity"] and
								metal.stock_uom == e_item["uom"]
							):
								key = (e_item["item_type"], e_item["uom"])
								
								if key not in aggregated_metal_items:
									aggregated_metal_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = metal.quantity * item.qty
								metal_rate = metal.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else metal.rate
								# making_amount=metal.making_amount
								metal_amount = (metal_rate * multiplied_qty)
								
								# Sum quantities and amounts
								aggregated_metal_items[key]["qty"] += multiplied_qty
								aggregated_metal_items[key]["amount"] += metal_amount

								# Calculate tax amount
								tax_rate_decimal = aggregated_metal_items[key]["tax_rate"] / 100
								aggregated_metal_items[key]["tax_amount"] += metal_amount * tax_rate_decimal

								aggregated_metal_items[key]["amount_with_tax"] = (
									aggregated_metal_items[key]["amount"] +
									aggregated_metal_items[key]["tax_amount"]
								)
								break
								
						for e_item in e_invoice_items:
							if (
								e_item["is_for_making"] and
								metal.metal_type == e_item["metal_type"] and
								metal.metal_touch == e_item["metal_purity"] and
								metal.stock_uom == e_item["uom"]
							):
								key = (e_item["item_type"], e_item["uom"])

								if key not in aggregated_metal_making_items:
									aggregated_metal_making_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": metal.making_rate,  # initial rate, will be overwritten with average later
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = metal.quantity * item.qty
								metal_making_amount = metal.making_rate * multiplied_qty
								aggregated_metal_making_items[key]["qty"] += multiplied_qty
								aggregated_metal_making_items[key]["amount"] += metal_making_amount

								tax_rate_decimal = aggregated_metal_making_items[key]["tax_rate"] / 100
								aggregated_metal_making_items[key]["tax_amount"] += metal_making_amount * tax_rate_decimal

								aggregated_metal_making_items[key]["amount_with_tax"] = (
										aggregated_metal_making_items[key]["amount"] +
									aggregated_metal_making_items[key]["tax_amount"]
								)
								break
					else:
						for e_item in e_invoice_items:
							
							if (
								e_item["is_for_labour"]
								# and metal.stock_uom == e_item["uom"]
								# and metal.metal_type == e_item["metal_type"]
								# and metal.metal_touch == e_item["metal_purity"]
							):
								key = (e_item["item_type"], e_item["uom"])
								if key not in aggregated_metal_labour_items:
									aggregated_metal_labour_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = metal.quantity * item.qty
								# metal_rate = metal.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else metal.making_rate
								metal_rate =metal.making_rate
								metal_amount = metal_rate * multiplied_qty

								aggregated_metal_labour_items[key]["qty"] += multiplied_qty
								aggregated_metal_labour_items[key]["amount"] += metal_amount
								tax_rate_decimal = aggregated_metal_labour_items[key]["tax_rate"] / 100
								aggregated_metal_labour_items[key]["tax_amount"] += metal_amount * tax_rate_decimal
								aggregated_metal_labour_items[key]["amount_with_tax"] = (
									aggregated_metal_labour_items[key]["amount"] +
									aggregated_metal_labour_items[key]["tax_amount"]
								)
								
						

				for diamond in bom_doc.diamond_detail:
					if not diamond.is_customer_item:
						for e_item in e_invoice_items:
							if (
								e_item["is_for_diamond"]
								and e_item["diamond_type"] == diamond.diamond_type
								and e_item["uom"] == diamond.stock_uom
							):
								key = (e_item["item_type"], e_item["uom"])

								if key not in aggregated_diamond_items:
									aggregated_diamond_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = diamond.quantity * item.qty
								diamond_rate = diamond.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else diamond.total_diamond_rate
								diamond_amount = flt(diamond.diamond_rate_for_specified_quantity)

								aggregated_diamond_items[key]["qty"] += multiplied_qty
								aggregated_diamond_items[key]["amount"] += diamond_amount

								# Calculate average rate after accumulation
								if aggregated_diamond_items[key]["qty"] > 0:
									aggregated_diamond_items[key]["rate"] = aggregated_diamond_items[key]["amount"] / aggregated_diamond_items[key]["qty"]
								else:
									aggregated_diamond_items[key]["rate"] = 0

								tax_rate_decimal = aggregated_diamond_items[key]["tax_rate"] / 100
								aggregated_diamond_items[key]["tax_amount"] += diamond_amount * tax_rate_decimal

								aggregated_diamond_items[key]["amount_with_tax"] = (
									aggregated_diamond_items[key]["amount"] +
									aggregated_diamond_items[key]["tax_amount"]
								)
					else:
						for e_item in e_invoice_items:
							if (
								e_item["is_for_labour"]
								# and e_item["diamond_type"] == diamond.diamond_type
								# and e_item["uom"] == diamond.stock_uom
							):
								key = (e_item["item_type"], e_item["uom"])

								if key not in aggregated_metal_labour_items:
									aggregated_metal_labour_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = diamond.quantity * item.qty
								diamond_rate = diamond.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else diamond.total_diamond_rate
								diamond_amount = flt(diamond.diamond_rate_for_specified_quantity)

								aggregated_metal_labour_items[key]["qty"] += multiplied_qty/5
								aggregated_metal_labour_items[key]["amount"] += diamond_amount
								# Calculate average rate after accumulation
								if aggregated_metal_labour_items[key]["qty"] > 0:
									aggregated_metal_labour_items[key]["rate"] = aggregated_metal_labour_items[key]["amount"] / aggregated_metal_labour_items[key]["qty"]
								else:
									aggregated_metal_labour_items[key]["rate"] = 0

								tax_rate_decimal = aggregated_metal_labour_items[key]["tax_rate"] / 100
								aggregated_metal_labour_items[key]["tax_amount"] += diamond_amount * tax_rate_decimal

								aggregated_metal_labour_items[key]["amount_with_tax"] = (
									aggregated_metal_labour_items[key]["amount"] +
									aggregated_metal_labour_items[key]["tax_amount"]
								)		

				for gemstone in bom_doc.gemstone_detail:
					for e_item in e_invoice_items:
						if not gemstone.is_customer_item:
							if (
								e_item["is_for_gemstone"]
								and e_item["uom"] == gemstone.stock_uom
							):
								key = (e_item["item_type"], e_item["uom"])

								if key not in aggregated_gemstone_items:
									aggregated_gemstone_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": gemstone.total_gemstone_rate,  # initial rate; average will be calculated later
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = gemstone.quantity * item.qty
								gemstone_rate = gemstone.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else gemstone.total_gemstone_rate
								gemstone_amount = flt(gemstone.gemstone_rate_for_specified_quantity)

								aggregated_gemstone_items[key]["qty"] += multiplied_qty
								aggregated_gemstone_items[key]["amount"] += gemstone_amount

								# Calculate average rate after accumulation
								if aggregated_gemstone_items[key]["qty"] > 0:
									aggregated_gemstone_items[key]["rate"] = aggregated_gemstone_items[key]["amount"] / aggregated_gemstone_items[key]["qty"]
								else:
									aggregated_gemstone_items[key]["rate"] = 0

								tax_rate_decimal = aggregated_gemstone_items[key]["tax_rate"] / 100
								aggregated_gemstone_items[key]["tax_amount"] += gemstone_amount * tax_rate_decimal

								aggregated_gemstone_items[key]["amount_with_tax"] = (
									aggregated_gemstone_items[key]["amount"] +
									aggregated_gemstone_items[key]["tax_amount"]
								)
						else:
							if (
							e_item["is_for_labour"]
							and e_item["uom"] == gemstone.stock_uom
						):
								key = (e_item["item_type"], e_item["uom"])

								if key not in aggregated_metal_labour_items:
									aggregated_metal_labour_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": gemstone.total_gemstone_rate,  # initial rate; average will be calculated later
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = gemstone.quantity * item.qty
								gemstone_rate = gemstone.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else gemstone.total_gemstone_rate
								gemstone_amount = flt(gemstone.gemstone_rate_for_specified_quantity)

								aggregated_metal_labour_items[key]["qty"] += multiplied_qty/5
								aggregated_metal_labour_items[key]["amount"] += gemstone_amount
								# Calculate average rate after accumulation
								if aggregated_metal_labour_items[key]["qty"] > 0:
									aggregated_metal_labour_items[key]["rate"] = aggregated_metal_labour_items[key]["amount"] / aggregated_metal_labour_items[key]["qty"]
								else:
									aggregated_metal_labour_items[key]["rate"] = 0

								tax_rate_decimal = aggregated_metal_labour_items[key]["tax_rate"] / 100
								aggregated_metal_labour_items[key]["tax_amount"] += gemstone_amount * tax_rate_decimal

								aggregated_metal_labour_items[key]["amount_with_tax"] = (
									aggregated_metal_labour_items[key]["amount"] +
									aggregated_metal_labour_items[key]["tax_amount"]
								)
				for finding in bom_doc.finding_detail:
					if not finding.is_customer_item:
						finding_handled = False
						for e_item in e_invoice_items:
							if (e_item["is_for_finding"] and e_item["metal_type"] == finding.metal_type and e_item["metal_purity"] == finding.metal_touch and e_item["uom"] == finding.stock_uom and e_item["finding_category"] == finding.finding_category):
								finding_handled = True
								key = (e_item["item_type"], e_item["uom"])
								if key not in aggregated_finding_items:
									aggregated_finding_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}
								multiplied_qty = finding.quantity * item.qty
								making_amount = finding.making_amount
								finding_rate = 0 
								if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009":
									finding_rate = finding.se_rate 
								elif self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009":
									finding_rate = finding.se_rate
								elif self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009":
									finding_rate = finding.se_rate 
								finding_making_amount = (finding.rate * multiplied_qty)
								aggregated_finding_items[key]["qty"] += multiplied_qty
								aggregated_finding_items[key]["amount"] += finding_making_amount
								aggregated_finding_items[key]["rate"] = finding_rate
								
								tax_rate_decimal = aggregated_finding_items[key]["tax_rate"] / 100
								aggregated_finding_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal

								aggregated_finding_items[key]["amount_with_tax"] = (
									aggregated_finding_items[key]["amount"] +
									aggregated_finding_items[key]["tax_amount"]
								)
								break

						if not finding_handled:
							for e_item in e_invoice_items:
								if (e_item["is_for_metal"] and finding.metal_type == e_item["metal_type"] and finding.metal_touch == e_item["metal_purity"] and finding.stock_uom == e_item["uom"] and e_item["finding_category"] is None):
									key = (e_item["item_type"], e_item["uom"])
									if key not in aggregated_metal_items:
										aggregated_metal_items[key] = {
											"item_code": e_item["item_type"],
											"item_name": e_item["item_type"],
											"uom": e_item["uom"],
											"qty": 0,
											"amount": 0,
											"tax_rate": e_item["tax_rate"],
											"tax_amount": 0,
											"amount_with_tax": 0,
											"delivery_date": self.delivery_date,
											"rate": 0
										}
									
									finding_rate = finding.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else finding.rate
									multiplied_qty = finding.quantity * item.qty
									making_amount = finding.making_amount
									finding_making_amount = (finding.rate * multiplied_qty)
									
									aggregated_metal_items[key]["qty"] += multiplied_qty
									aggregated_metal_items[key]["amount"] += finding_making_amount
									aggregated_metal_items[key]["rate"] = finding_rate
									
									tax_rate_decimal = aggregated_metal_items[key]["tax_rate"] / 100
									aggregated_metal_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal
									aggregated_metal_items[key]["amount_with_tax"] = (
										aggregated_metal_items[key]["amount"] + 
										aggregated_metal_items[key]["tax_amount"]
									)
									break

						
						finding_making_handled = False
						for e_item in e_invoice_items:
							if (e_item["is_for_finding_making"] and e_item["metal_type"] == finding.metal_type and e_item["metal_purity"] == finding.metal_touch and e_item["uom"] == finding.stock_uom and e_item["finding_category"] == finding.finding_category):
								finding_making_handled = True
								key = (e_item["item_type"], e_item["uom"])
								if key not in aggregated_finding_making_items:
									aggregated_finding_making_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": finding.making_rate,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}
								
								multiplied_qty = finding.quantity * item.qty
								making_amount = finding.making_amount
								finding_making_amount = (finding.making_rate * multiplied_qty)
								
								aggregated_finding_making_items[key]["qty"] += multiplied_qty
								aggregated_finding_making_items[key]["amount"] += finding_making_amount
								if aggregated_finding_making_items[key]["qty"] > 0:
									aggregated_finding_making_items[key]["rate"] = aggregated_finding_making_items[key]["amount"] / aggregated_finding_making_items[key]["qty"]
								else:
									aggregated_finding_making_items[key]["rate"] = 0
								
								tax_rate_decimal = aggregated_finding_making_items[key]["tax_rate"] / 100
								aggregated_finding_making_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal
								aggregated_finding_making_items[key]["amount_with_tax"] = (
									aggregated_finding_making_items[key]["amount"] +
									aggregated_finding_making_items[key]["tax_amount"]
								)
								break
						
						if not finding_making_handled:
							for e_item in e_invoice_items:
								if (e_item["is_for_making"] and e_item["metal_type"] == finding.metal_type and e_item["metal_purity"] == finding.metal_touch and e_item["uom"] == finding.stock_uom):
									key = (e_item["item_type"], e_item["uom"])
									if key not in aggregated_metal_making_items:
										aggregated_metal_making_items[key] = {
											"item_code": e_item["item_type"],
											"item_name": e_item["item_type"],
											"uom": e_item["uom"],
											"qty": 0,
											"rate": finding.making_rate,
											"amount": 0,
											"tax_rate": e_item["tax_rate"],
											"tax_amount": 0,
											"amount_with_tax": 0,
											"delivery_date": self.delivery_date
										}
									
									multiplied_qty = finding.quantity * item.qty
									making_amount = finding.making_amount
									finding_making_amount = (finding.making_rate * multiplied_qty)
									aggregated_metal_making_items[key]["qty"] += multiplied_qty
									aggregated_metal_making_items[key]["amount"] += finding_making_amount

									if aggregated_metal_making_items[key]["qty"] > 0:
										aggregated_metal_making_items[key]["rate"] = aggregated_metal_making_items[key]["amount"] / aggregated_metal_making_items[key]["qty"]
									else:
										aggregated_metal_making_items[key]["rate"] = 0
									
									tax_rate_decimal = aggregated_metal_making_items[key]["tax_rate"] / 100
									aggregated_metal_making_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal
									aggregated_metal_making_items[key]["amount_with_tax"] = (
										aggregated_metal_making_items[key]["amount"] +
										aggregated_metal_making_items[key]["tax_amount"]
									)
									break
					else:
						for e_item in e_invoice_items:
							if (e_item["is_for_labour"] ):
								key = (e_item["item_type"], e_item["uom"])
								if key not in aggregated_metal_labour_items:
									aggregated_metal_labour_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": finding.making_rate,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}
								
								multiplied_qty = finding.quantity * item.qty
								making_amount = finding.making_amount
								finding_making_amount = (finding.making_rate * multiplied_qty)
								aggregated_metal_labour_items[key]["qty"] += multiplied_qty
								aggregated_metal_labour_items[key]["amount"] += finding_making_amount

								if aggregated_metal_labour_items[key]["qty"] > 0:
									aggregated_metal_labour_items[key]["rate"] = aggregated_metal_labour_items[key]["amount"] / aggregated_metal_labour_items[key]["qty"]
								else:
									aggregated_metal_labour_items[key]["rate"] = 0
								
								tax_rate_decimal = aggregated_metal_labour_items[key]["tax_rate"] / 100
								aggregated_metal_labour_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal
								aggregated_metal_labour_items[key]["amount_with_tax"] = (
									aggregated_metal_labour_items[key]["amount"] +
									aggregated_metal_labour_items[key]["tax_amount"]
								)
								break

		
		# After aggregation, calculate average rate = total amount / total qty per key
		for key, val in aggregated_metal_items.items():
			if val["qty"] > 0:
				
				average_rate = val["amount"] / val["qty"]
			else:
				average_rate = 0
			val["rate"] = average_rate
			self.append("custom_invoice_item", val)

		for key, val in aggregated_hallmarking_items.items():
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			self.append("custom_invoice_item", val)
		

		for key, val in aggregated_metal_labour_items.items():
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			val["qty"] = round(val["qty"],2)
			self.append("custom_invoice_item", val)
		
		for key, val in aggregated_metal_making_items.items():
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			self.append("custom_invoice_item", val)
		
		for key, val in aggregated_diamond_items.items():
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			self.append("custom_invoice_item", val)

		for key, val in aggregated_gemstone_items.items():
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			self.append("custom_invoice_item", val)
		
		for key, val in aggregated_finding_items.items():
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			self.append("custom_invoice_item", val)

	
		for key, val in aggregated_finding_making_items.items():
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			self.append("custom_invoice_item", val)




def validate_quotation_item(self):
	if not self.custom_invoice_item:
		for row in self.items:
			if row.prevdoc_docname:
				quotation_id = row.prevdoc_docname
				invoice_items = frappe.get_all(
					'Quotation E Invoice Item',
					filters={'parent': quotation_id},  
					fields=['item_code', 'item_name', 'uom', 'qty', 'rate', 'amount']
				)
				if invoice_items:
					for invoice_item in invoice_items:
						self.append('custom_invoice_item', {
							'item_code': invoice_item.item_code,
							'item_name': invoice_item.item_name,
							'uom': invoice_item.uom,
							'qty': invoice_item.qty,
							'rate': invoice_item.rate,
							'amount': invoice_item.amount
						})



def validate_sales_type(self):
	for r in self.items:
	# 	if r.prevdoc_docname:
	# 		quotation_sales_type = frappe.db.get_value('Quotation', r.prevdoc_docname, 'custom_sales_type')
	# 		if quotation_sales_type:  
	# 			self.sales_type = quotation_sales_type
	# 	if self.company == "Gurukrupa Export Private Limited":
	# 		# Throw only if BOTH are missing
		if not r.prevdoc_docname and not r.custom_customer_approval:
			pass
			# frappe.msgprint(
			# 	_("Row {0} : Sales Order can be created only from Quotation or Customer Approval for this Company").format(r.idx)
			# )
	if not self.sales_type :
		frappe.throw("Sales Type is mandatory.")
	# if not self.gold_rate_with_gst and self.company != 'Sadguru Diamond':
	# 	frappe.throw("Metal rate  with GST is mandatory.")



import json
@frappe.whitelist()

def make_sales_order_batch(sales_orders, target_doc=None):

	if isinstance(sales_orders, str):
		sales_orders = json.loads(sales_orders)

	if target_doc:
		if isinstance(target_doc, str):
			target_doc = json.loads(target_doc)

		target_doc = frappe.get_doc(target_doc)
	else:
		target_doc = frappe.new_doc("Sales Order")


	target_doc.items = []

	for so_name in sales_orders:
		so = frappe.db.get_value("Sales Order", so_name, "*", as_dict=True)
		if not so:
			continue
		target_doc.custom_diamond_quality = so.custom_diamond_quality
		target_doc.order_type = so.order_type
		target_doc.custom_parent_sales_order = so.name
		items = frappe.get_all(
			"Sales Order Item",
			filters={"parent": so_name},
			fields="*"
		)
		
		
		
		for it in items:
			snc_list = frappe.db.get_list("Serial Number Creator", 
				filters={"sales_order_id": so_name}, 
				fields=["name"]
			)
			
			stock_entries = []
			for snc in snc_list:
				stock_entry = frappe.db.get_value("Stock Entry", 
					{"custom_serial_number_creator": snc.name}, "name")
				if stock_entry:
					stock_entries.append(stock_entry)
			
			available_serials = []
			for stock_entry in stock_entries:
				serial_no = frappe.db.sql(f"""
					SELECT sed.serial_no, sed.item_code
					FROM `tabStock Entry Detail` sed
					WHERE sed.parent = '{stock_entry}'
					AND sed.item_code = '{it.item_code}'
					ORDER BY sed.idx DESC
					LIMIT 1
				""", as_dict=1)
				
				if serial_no and serial_no[0]['item_code'] == it.item_code:
					available_serials.append(serial_no[0]['serial_no'])
			
			if not available_serials:
				continue
			
			serial_count = 0
			for s_no in available_serials:
				if serial_count < it.qty:
					target_doc.append("items", {
						"item_code": it.item_code,
						"item_name": it.item_name,
						"serial_no": s_no,
						"bom": frappe.db.get_value("Serial No", s_no, "custom_bom_no"),
						"diamond_quality": so.custom_diamond_quality,
						"description": it.description,
						"qty": 1,
						"rate": it.rate,
						"warehouse": it.warehouse,
						"against_sales_order": so_name,
						"uom": it.uom
					})
					serial_count += 1
				else:
					break


	first_so = frappe.db.get_value("Sales Order", sales_orders[0], "*", as_dict=True)


	return target_doc
